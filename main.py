#!/usr/bin/python3
# -*- coding: utf-8 -*-

import os

from datefinder import find_dates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Border, Side

from employee_arrival_report import employee_arrival_report, title


def cell_framing(row, column, result_sheet):
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    result_sheet.cell(row=row, column=column).border = THIN_BORDER

work_directory = os.getcwd()


def main():
    # Создаём файл для записи результата
    wb = Workbook()
    result_sheet = wb.active

    # Устанавливаем заголовки столбцов
    result_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title_cell = result_sheet.cell(row=1, column=1)
    cell_framing(row=1, column=1, result_sheet=result_sheet)
    title_cell.value = title
    title_cell.font = Font(bold=True)

    COLUMNS_NAMES = ['Фамилия', 'Имя', 'Отчество', 'Должность', 'Приход']
    for index, column_name in enumerate(COLUMNS_NAMES):
        cell = result_sheet.cell(row=2, column=index+1)
        cell_framing(row=2, column=index+1, result_sheet=result_sheet)
        cell.value = column_name
        cell.font = Font(bold=True)
        cell.fill = PatternFill(bgColor='FFFF00', fill_type='gray0625')

    # Устанавливаем ширину столбцов
    COLUMNS = [('A', 12), ('B', 10), ('C', 14), ('D', 20), ('E', 10)]
    for column, width in COLUMNS:
        result_sheet.column_dimensions[column].width = width

    # Записываем данные таблицы
    for row, employee in enumerate(employee_arrival_report()):
        row = row+3
        for column, value in enumerate(employee):
            column = column+1
            cell = result_sheet.cell(row=row, column=column)
            cell_framing(row=row, column=column, result_sheet=result_sheet)
            cell.value = value
            if employee[4] > '9:00:00':
                cell.fill = PatternFill(bgColor='FF6D6D', fill_type='gray0625')

    # Из заголовка таблицы получаем название для файла
    dates = list(find_dates(title))
    fd = dates[0].timetuple() # форматированная дата
    DayMonthYear = [fd.tm_mday, fd.tm_mon, fd.tm_year]
    date = [str(item) for item in DayMonthYear]
    new_file_name = '.'.join(date) # дата = название файла

    wb.save(work_directory + f'/{new_file_name}.xlsx')


if __name__ == '__main__':
    input()
    main()
    input()