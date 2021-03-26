import openpyxl
import PySimpleGUI as sg

filename = sg.popup_get_file(
    message='Загрузите excel файл для обработки',
    size=(80, 100),
)

# Открывает excel файл, в котором лежит таблица с данными
# wb_with_table = openpyxl.load_workbook('test_table.xlsx')
wb_with_table = openpyxl.load_workbook(filename=filename)
work_sheet = wb_with_table.active

# Заголовок таблицы
title = work_sheet.cell(row=1, column=1).value


def get_cell_value(index, column_num):
    return work_sheet.cell(row=index, column=column_num).value


def employee_arrival_report():
    row_for_write = 3
    BAD_RESULT = [None, '-']
    # Сотрудники со сменным графиком, не входят в список контролируемых
    SURNAMES_FOR_DELETE = ['Иванов', 'Петрова', 'Шубейкин', 'Петрушкин']
    result = []
    # Берём данные из таблицы (количество строк)
    rows_amount = work_sheet.max_row

    for index in range(3, rows_amount-2):
        time_in = get_cell_value(index, 8)
        surname = get_cell_value(index, 1)
        if time_in not in BAD_RESULT and surname not in SURNAMES_FOR_DELETE:
            name = get_cell_value(index, 2)
            middle_name = get_cell_value(index, 3)
            positon = get_cell_value(index, 7)

            result.append((surname, name, middle_name, positon, time_in))
            row_for_write += 1

    result = sorted(result, key=lambda x: x[4])
    return result
